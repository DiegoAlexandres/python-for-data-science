#%%
# Nosso objetivo é criar um novo arquivo Excel que contenha:
# Uma nova aba chamada "Contas a Receber com Nomes", que seja uma cópia da fContasReceber, mas com uma coluna extra contendo o nome do cliente correspondente.
# Uma segunda nova aba chamada "Resumo por Cliente", que calcule e mostre o valor total a receber de cada cliente.

import openpyxl
from openpyxl.utils import get_column_letter

#%%
# --- ETAPA 1: Carregar o arquivo e as planilhas ---
# Carrega o arquivo Excel 'base_de_dados.xlsx' para a memória.
workbook = openpyxl.load_workbook('base_de_dados.xlsx')

#%%
# Seleciona as planilhas com as quais vamos trabalhar.
sheet_clientes = workbook['dClientes']
sheet_contas_receber = workbook['fContasReceber']

#%%
# --- ETAPA 2: Ler os dados dos clientes e criar um mapa de referência ---
# Cria um dicionário para armazenar os nomes dos clientes usando o código como chave.
# Isso funciona como um PROCV (VLOOKUP) super rápido.
clientes_mapa = {}
# Itera sobre as linhas da planilha de clientes, começando da linha 2 para pular o cabeçalho.
for linha in sheet_clientes.iter_rows(min_row=2, values_only=True):
    cod_cliente = linha[0]
    nome_cliente = linha[1]
    clientes_mapa[cod_cliente] = nome_cliente

#%%
# --- ETAPA 3: Processar as contas a receber e enriquecer os dados ---
# Cria uma lista para guardar os novos dados que terão o nome do cliente.
contas_receber_com_nome = []

# Adiciona o novo cabeçalho à nossa lista.
cabecalho_novo = [cell.value for cell in sheet_contas_receber[1]] + ["Nome Cliente"]
contas_receber_com_nome.append(cabecalho_novo)

#%%
# Itera sobre as contas a receber, pulando o cabeçalho.
for linha in sheet_contas_receber.iter_rows(min_row=2, values_only=True):
    cod_cliente_da_conta = linha[5] # Coluna 'CodCliente' está na 6ª posição (índice 5)
    
    # Busca o nome do cliente no nosso mapa. Se não encontrar, usa um valor padrão.
    nome_cliente = clientes_mapa.get(cod_cliente_da_conta, "CLIENTE NÃO ENCONTRADO")
    
    # Adiciona o nome do cliente ao final da linha original.
    linha_nova = list(linha) + [nome_cliente]
    contas_receber_com_nome.append(linha_nova)

#%%
# --- ETAPA 4: Calcular o resumo de valores por cliente ---
total_por_cliente = {}
# Itera novamente sobre as contas a receber para somar os valores.
for linha in sheet_contas_receber.iter_rows(min_row=2, values_only=True):
    cod_cliente_da_conta = linha[5]
    valor_str = str(linha[6]).replace(',', '.') # Pega o valor e substitui vírgula por ponto.
    
    try:
        valor = float(valor_str)
        nome_cliente = clientes_mapa.get(cod_cliente_da_conta, "CLIENTE NÃO ENCONTRADO")
        
        # Adiciona e soma o valor para o cliente correspondente.
        total_por_cliente[nome_cliente] = total_por_cliente.get(nome_cliente, 0) + valor
    except (ValueError, TypeError):
        print(f"Não foi possível converter o valor '{linha[6]}' para número na linha.")

#%%
# --- ETAPA 5: Criar o novo arquivo Excel com os resultados ---
# Cria um novo Workbook (arquivo Excel) em memória.
novo_workbook = openpyxl.Workbook()

#%%
# Remove a planilha padrão que é criada.
novo_workbook.remove(novo_workbook.active)

#%%
# Cria a primeira planilha para o relatório detalhado.
sheet_relatorio = novo_workbook.create_sheet("Contas a Receber com Nomes")
for linha in contas_receber_com_nome:
    sheet_relatorio.append(linha) # Adiciona cada linha de dados.

#%%
# Cria a segunda planilha para o resumo.
sheet_resumo = novo_workbook.create_sheet("Resumo por Cliente")
sheet_resumo.append(["Nome Cliente", "Valor Total Recebido"]) # Adiciona o cabeçalho.
for nome, total in total_por_cliente.items():
    sheet_resumo.append([nome, total]) # Adiciona os dados do resumo.

#%%    
# --- ETAPA 6: Salvar o novo arquivo ---
# Salva o arquivo Excel com o nome 'relatorio_final.xlsx'.
novo_workbook.save("relatorio_final.xlsx")

print("Relatório 'relatorio_final.xlsx' gerado com sucesso!")
